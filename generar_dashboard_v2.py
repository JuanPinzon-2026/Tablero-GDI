# generar_dashboard_v2.py
# Inyecta datos cifrados en index_v2.html (Tablero GDI v2 - Alpine.js)
# Misma logica de lectura y cifrado que generar_dashboard.py

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl, json, re, os, platform
from datetime import datetime, date

# -- Clave de cifrado ---------------------------------------------------------
CLAVE_TABLERO = "GDI2026"

# -- Cifrado AES-256-GCM ------------------------------------------------------
def _asegurar_cryptography():
    try:
        import cryptography  # noqa
    except ImportError:
        import subprocess
        print("  Instalando 'cryptography'...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "cryptography"],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def cifrar_datos(texto, clave):
    import os as _os, base64
    _asegurar_cryptography()
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.backends import default_backend
    salt = _os.urandom(16)
    iv   = _os.urandom(12)
    kdf  = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt,
                      iterations=100000, backend=default_backend())
    key  = kdf.derive(clave.encode('utf-8'))
    aesgcm     = AESGCM(key)
    ciphertext = aesgcm.encrypt(iv, texto.encode('utf-8'), None)
    return base64.b64encode(salt + iv + ciphertext).decode('utf-8')

# -- Rutas --------------------------------------------------------------------
BASE      = os.path.dirname(os.path.abspath(__file__))
DASHBOARD = os.path.join(BASE, "index_v2.html")

if platform.system() == "Windows":
    ORDENES_DIR = os.path.join(os.environ.get('USERPROFILE', os.path.expanduser('~')),
                               'Software Broker',
                               'Nathalia Moreno - Ordenes incidentadas 2026')
else:
    ORDENES_DIR = "/sessions/affectionate-wonderful-fermat/mnt/Nathalia Moreno - Ordenes incidentadas 2026"

EXCELS_MAIN = [
    os.path.join(ORDENES_DIR, "Ordenes con novedad Junio - dic.xlsx"),
]
EXCELS_SS = [
    os.path.join(ORDENES_DIR, "Ordenes con novedad Enero (1).xlsx"),
    os.path.join(ORDENES_DIR, "Ordenes con novedad Junio - dic.xlsx"),
]
IXC_EXCEL = os.path.join(BASE, "IXComercio.xlsx")

# -- Mapeo de columnas --------------------------------------------------------
PRIORITY_SHEETS = ["novedad", "ordenes", "data", "datos", "incidentadas"]
COLUMN_MAP = {
    "fo":    ["fecha de orden", "fecha orden", "fecha_orden", "fo", "f. orden", "fecha oc"],
    "fn":    ["fecha de notificacion jira", "fecha notificacion", "fecha novedad", "fecha_novedad", "fn", "f. novedad", "fecha nov", "fecha de notificacion"],
    "com":   ["comentario continuidad", "rta continuidad", "comentario", "continuidad", "com", "comentarios"],
    "est":   ["estado caso", "error", "estado", "error / estado", "error/estado", "est", "descripcion", "descripcion", "motivo"],
    "pen":   ["pendiente por:", "pendiente por", "pendiente", "pen", "pendiente x", "pend"],
    "marca": ["marca", "brand", "cliente", "tienda"],
    "pais":  ["marca pais", "pais", "pais", "marca_pais", "country"],
    "dup":   ["duplicado", "dup", "estado duplicado", "en revision"],
    "ops":   ["accion ops", "accion_ops", "ops", "accion", "mitigacion", "estado de revision", "accion de ops"],
    "ticket": ["ticket", "jira", "no jira", "numero jira", "ithd", "no ticket", "jira ticket"],
    "venta": ["venta", "nro venta", "numero venta", "no. venta", "orden de venta", "orden", "order", "nro orden", "numero de orden"],
    "prov":  ["proveedor", "provider", "prov"],
}

def normalizar(s):
    if not isinstance(s, str): return ""
    s = s.lower().strip()
    for a, b in [("\xe1","a"),("\xe9","e"),("\xed","i"),("\xf3","o"),("\xfa","u"),("\xf1","n")]:
        s = s.replace(a, b)
    return s

def detectar_columnas(headers):
    norm_headers = [(i, normalizar(str(h))) for i, h in enumerate(headers) if h]
    mapping = {}
    for campo, candidatos in COLUMN_MAP.items():
        best_col, best_score = None, -1
        for i, nh in norm_headers:
            for c in candidatos:
                if c == nh:          score = len(c) * 100
                elif c in nh:        score = len(c)
                elif nh in c:        score = len(nh)
                else:                continue
                if score > best_score: best_score = score; best_col = i
        if best_col is not None: mapping[campo] = best_col
    requeridos = {"fo", "fn", "com", "est", "pen", "marca"}
    faltantes = requeridos - set(mapping.keys())
    if faltantes:
        print(f"  WARNING Columnas no encontradas: {faltantes}")
    return mapping

def fmt_fecha(val):
    if isinstance(val, (datetime, date)): return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        v = val.strip()
        m = re.match(r'^(\d{2})[/\-](\d{2})[/\-](\d{4})$', v)
        if m: return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        return v[:10]
    return ""

def leer_excel(path):
    print(f"  Leyendo: {os.path.basename(path)}")
    if not os.path.exists(path):
        print(f"  ERROR No existe: {path}")
        return []
    import tempfile, shutil
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        shutil.copy2(path, tmp.name)
    except Exception as e:
        print(f"  ERROR copiando: {e}"); return []
    try:
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo Excel: {e}"); os.remove(tmp.name); return []
    records = []
    sheets_to_read = wb.sheetnames
    priority = [s for s in wb.sheetnames if normalizar(s) in PRIORITY_SHEETS]
    if priority:
        sheets_to_read = priority
        print(f"  -> Hojas: {priority}")
    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        header_row, header_idx = None, 0
        for idx, row in enumerate(rows[:5]):
            if any(cell for cell in row if cell):
                header_row = row; header_idx = idx; break
        if header_row is None: continue
        col_map = detectar_columnas(header_row)
        if len(col_map) < 6:
            print(f"  WARNING Hoja '{sheet_name}': pocas columnas, omitiendo."); continue
        print(f"  OK Hoja '{sheet_name}': {len(col_map)} columnas")
        for row in rows[header_idx + 1:]:
            if not any(row): continue
            def get(campo):
                i2 = col_map.get(campo)
                if i2 is None or i2 >= len(row): return ""
                val = row[i2]
                return str(val).strip() if val is not None else ""
            fo = fmt_fecha(row[col_map["fo"]] if "fo" in col_map and col_map["fo"] < len(row) else "")
            fn = fmt_fecha(row[col_map["fn"]] if "fn" in col_map and col_map["fn"] < len(row) else "")
            if not fo and not fn and not get("ops"): continue
            records.append({
                "fo": fo, "fn": fn, "com": get("com"), "est": get("est"),
                "pen": get("pen"), "marca": get("marca"), "pais": get("pais"),
                "dup": get("dup"), "ops": get("ops"), "venta": get("venta"),
                "ticket": get("ticket"), "prov": get("prov").strip().title()
            })
    wb.close(); os.remove(tmp.name)
    print(f"  -> {len(records)} registros")
    return records

def leer_ixc_excel(path):
    if not os.path.exists(path):
        print(f"  INFO IXC Excel no encontrado (opcional)"); return {}
    print(f"  Leyendo IXC: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active; data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_raw = row[0] if len(row) > 0 else None
        total     = row[4] if len(row) > 4 else None
        if fecha_raw is None or total is None: continue
        iso = ""
        if isinstance(fecha_raw, (datetime, date)):
            iso = fecha_raw.strftime("%Y-%m-%d")
        elif isinstance(fecha_raw, str):
            m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", fecha_raw.strip())
            if m: iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        if iso and re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
            try: data[iso] = int(total)
            except (ValueError, TypeError): pass
    wb.close()
    print(f"  -> {len(data)} fechas IXC")
    return data

def es_sin_stock(r):
    return any("sin stock" in (r.get(k, "")).lower() for k in ("com", "ops", "est"))

def reemplazar_string(content, variable, value):
    m = re.search(rf'(?<!\w){re.escape(variable)}\s*=\s*\'[^\']*\'', content)
    if not m:
        print(f"  ERROR No se encontro '{variable}' en index_v2.html")
        return None
    return content[:m.start()] + f"{variable}='{value}'" + content[m.end():]

def actualizar_html(records_main, records_ss, ixc_data):
    if not os.path.exists(DASHBOARD):
        print(f"ERROR No se encontro: {DASHBOARD}"); return False
    with open(DASHBOARD, "r", encoding="utf-8") as f:
        content = f.read()

    print("  Cifrando datos...")
    enc_main = cifrar_datos(json.dumps(records_main,  ensure_ascii=False, separators=(",",":")), CLAVE_TABLERO)
    enc_ss   = cifrar_datos(json.dumps(records_ss,    ensure_ascii=False, separators=(",",":")), CLAVE_TABLERO)
    enc_ixc  = cifrar_datos(json.dumps(ixc_data or {}, ensure_ascii=False, separators=(",",":")), CLAVE_TABLERO)
    print(f"  OK: MAIN={len(enc_main)}c, SS={len(enc_ss)}c, IXC={len(enc_ixc)}c")

    content = reemplazar_string(content, "_RECORDS_ENC",    enc_main)
    if content is None: return False
    content = reemplazar_string(content, "_RECORDS_SS_ENC", enc_ss)
    if content is None: return False
    content = reemplazar_string(content, "_IXC_ENC",        enc_ixc)
    if content is None: return False

    import os as _os
    with open(DASHBOARD, "w", encoding="utf-8") as f:
        f.write(content); f.flush(); _os.fsync(f.fileno())
    return True

# -- Main ---------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{'='*50}")
    print(f"  Tablero GDI v2 - Generador de Dashboard")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}\n")

    records_main = []
    for path in EXCELS_MAIN:
        records_main.extend(leer_excel(path))
    print(f"\nRECORDS (principal): {len(records_main)} registros\n")

    records_ss_raw = []
    for path in EXCELS_SS:
        records_ss_raw.extend(leer_excel(path))
    records_ss = [r for r in records_ss_raw if es_sin_stock(r)]
    print(f"RECORDS_SS (sin stock): {len(records_ss)} de {len(records_ss_raw)}\n")

    if not records_main:
        print("WARNING Sin registros en archivo principal."); exit(1)

    ixc_data = leer_ixc_excel(IXC_EXCEL)

    ok = actualizar_html(records_main, records_ss, ixc_data)
    if not ok: exit(1)

    print("\nOK index_v2.html actualizado.")
