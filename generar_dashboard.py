# generar_dashboard.py
# Lee los Excel de ordenes y actualiza el index.html del Tablero GDI

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
import json
import re
import os
import platform
from datetime import datetime, date

# -- Clave de cifrado ---------------------------------------------------------
# Cambia esta clave antes de publicar. Compártela sólo con el equipo GDI.
CLAVE_TABLERO = "GDI2026"

# -- Cifrado AES-256-GCM ------------------------------------------------------
def _asegurar_cryptography():
    """Instala 'cryptography' automáticamente si no está disponible."""
    try:
        import cryptography  # noqa
    except ImportError:
        import subprocess, sys
        print("  Libreria 'cryptography' no encontrada. Instalando automaticamente...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "cryptography"],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("  Libreria instalada OK")

def cifrar_datos(texto, clave):
    """Cifra texto con AES-256-GCM + PBKDF2-SHA256. Compatible con Web Crypto API del navegador."""
    import os as _os, base64
    _asegurar_cryptography()
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.backends import default_backend
    salt = _os.urandom(16)
    iv   = _os.urandom(12)   # 96 bits — recomendado para GCM
    kdf  = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt,
                      iterations=100000, backend=default_backend())
    key  = kdf.derive(clave.encode('utf-8'))
    aesgcm     = AESGCM(key)
    ciphertext = aesgcm.encrypt(iv, texto.encode('utf-8'), None)
    # Formato: base64( salt[16] + iv[12] + ciphertext+tag )
    return base64.b64encode(salt + iv + ciphertext).decode('utf-8')

# -- Rutas --------------------------------------------------------------------
BASE = os.path.dirname(os.path.abspath(__file__))
DASHBOARD = os.path.join(BASE, "index.html")

# Detectar si corre en Linux (sandbox) o Windows
if platform.system() == "Windows":
    ORDENES_DIR = r"C:\Users\jpinz390\Software Broker\Nathalia Moreno - Ordenes incidentadas 2026"
else:
    # Ruta de montaje en sandbox Linux
    ORDENES_DIR = "/sessions/affectionate-wonderful-fermat/mnt/Nathalia Moreno - Ordenes incidentadas 2026"

EXCELS_MAIN = [
    os.path.join(ORDENES_DIR, "Ordenes con novedad Junio - dic.xlsx"),
]

# Excel IXC (Total ORs generadas en IX Comercio) — guardar con este nombre en la carpeta Dashboard
IXC_EXCEL = os.path.join(BASE, "IXComercio.xlsx")

EXCELS_SS = [
    os.path.join(ORDENES_DIR, "Ordenes con novedad Enero (1).xlsx"),
    os.path.join(ORDENES_DIR, "Ordenes con novedad Junio - dic.xlsx"),
]

# -- Mapeo de columnas --------------------------------------------------------
PRIORITY_SHEETS = ["novedad", "ordenes", "data", "datos", "incidentadas"]

COLUMN_MAP = {
    "fo":    ["fecha de orden", "fecha orden", "fecha_orden", "fo", "f. orden", "fecha oc"],
    "fn":    ["fecha de notificacion jira", "fecha notificacion", "fecha novedad", "fecha_novedad", "fn", "f. novedad", "fecha nov", "fecha de notificacion"],
    "com":   ["comentario continuidad", "rta continuidad", "comentario", "continuidad", "com", "comentarios"],
    "est":   ["estado caso", "error", "estado", "error / estado", "error/estado", "est", "descripcion", "descripcion", "motivo"],
    "pen":   ["pendiente por:", "pendiente por", "pendiente", "pen", "pendiente x", "pend"],
    "marca": ["marca", "brand", "cliente", "tienda"],
    "pais":  ["marca pais", "pais", "pais", "marca_pais", "country"],
    "dup":   ["duplicado", "dup", "estado duplicado", "en revision"],
    "ops":   ["accion ops", "accion_ops", "ops", "accion", "mitigacion", "estado de revision", "accion de ops"],
    "ticket": ["ticket", "jira", "no jira", "numero jira", "ithd", "no ticket", "jira ticket"],
    "venta": ["venta", "nro venta", "numero venta", "no. venta", "orden de venta", "orden", "order", "nro orden", "numero de orden"],
}

def normalizar(s):
    if not isinstance(s, str):
        return ""
    s = s.lower().strip()
    for a, b in [("\xe1","a"),("\xe9","e"),("\xed","i"),("\xf3","o"),("\xfa","u"),("\xf1","n")]:
        s = s.replace(a, b)
    return s

def detectar_columnas(headers):
    norm_headers = [(i, normalizar(str(h))) for i, h in enumerate(headers) if h]
    mapping = {}
    for campo, candidatos in COLUMN_MAP.items():
        best_col = None
        best_score = -1
        for i, nh in norm_headers:
            for c in candidatos:
                if c == nh:
                    score = len(c) * 100
                elif c in nh:
                    score = len(c)
                elif nh in c:
                    score = len(nh)
                else:
                    continue
                if score > best_score:
                    best_score = score
                    best_col = i
        if best_col is not None:
            mapping[campo] = best_col
    requeridos = {"fo", "fn", "com", "est", "pen", "marca"}
    faltantes = requeridos - set(mapping.keys())
    if faltantes:
        print(f"  WARNING Columnas no encontradas: {faltantes}")
        print(f"  Headers: {[h for _,h in norm_headers]}")
    return mapping

def fmt_fecha(val):
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        v = val.strip()
        # Convertir DD/MM/YYYY → YYYY-MM-DD
        m = re.match(r'^(\d{2})[/\-](\d{2})[/\-](\d{4})$', v)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        return v[:10]
    return ""

def leer_excel(path):
    print(f"  Leyendo: {os.path.basename(path)}")
    if not os.path.exists(path):
        print(f"  ERROR No existe: {path}")
        return []
    # Copiar a temp para evitar PermissionError si el archivo está abierto en Excel
    import tempfile, shutil
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        shutil.copy2(path, tmp.name)
    except Exception as e:
        print(f"  ERROR copiando archivo: {e}")
        return []
    try:
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo Excel: {e}")
        os.remove(tmp.name)
        return []
    records = []
    sheets_to_read = wb.sheetnames
    priority = [s for s in wb.sheetnames if normalizar(s) in PRIORITY_SHEETS]
    if priority:
        sheets_to_read = priority
        print(f"  -> Hojas: {priority}")
    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = None
        header_idx = 0
        for idx, row in enumerate(rows[:5]):
            if any(cell for cell in row if cell):
                header_row = row
                header_idx = idx
                break
        if header_row is None:
            continue
        col_map = detectar_columnas(header_row)
        if len(col_map) < 6:
            print(f"  WARNING Hoja '{sheet_name}': pocas columnas ({len(col_map)}), omitiendo.")
            continue
        print(f"  OK Hoja '{sheet_name}': {len(col_map)} columnas")
        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            def get(campo):
                i2 = col_map.get(campo)
                if i2 is None or i2 >= len(row):
                    return ""
                val = row[i2]
                return str(val).strip() if val is not None else ""
            fo = fmt_fecha(row[col_map["fo"]] if "fo" in col_map and col_map["fo"] < len(row) else "")
            fn = fmt_fecha(row[col_map["fn"]] if "fn" in col_map and col_map["fn"] < len(row) else "")
            ops_val = get("ops")
            if not fo and not fn and not ops_val:
                continue
            records.append({"fo": fo, "fn": fn, "com": get("com"), "est": get("est"),
                            "pen": get("pen"), "marca": get("marca"), "pais": get("pais"),
                            "dup": get("dup"), "ops": ops_val, "venta": get("venta"), "ticket": get("ticket")})
    wb.close()
    os.remove(tmp.name)
    print(f"  -> {len(records)} registros")
    return records

def reemplazar_string(content, variable, value):
    """Reemplaza VAR="..." con el nuevo valor cifrado (base64)."""
    m = re.search(rf'(?<!\w){re.escape(variable)}\s*=\s*"[^"]*"', content)
    if not m:
        print(f"ERROR No se encontro '{variable}' en index.html")
        return None
    return content[:m.start()] + f'{variable}="{value}"' + content[m.end():]

def reemplazar_array(content, variable, records):
    m = re.search(rf'(?<!\w){variable}\s*=\s*\[', content)
    if not m:
        print(f"ERROR No se encontro '{variable}' en index.html")
        return None
    start_bracket = m.end() - 1
    depth = 0
    in_string = False
    end_bracket = -1
    i = start_bracket
    while i < len(content):
        c = content[i]
        if in_string:
            if c == '\\':
                i += 2
                continue
            elif c == '"':
                in_string = False
        else:
            if c == '"':
                in_string = True
            elif c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    end_bracket = i
                    break
        i += 1
    if end_bracket == -1:
        print(f"ERROR No se encontro cierre de array para '{variable}'")
        return None
    new_json = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    return content[:start_bracket] + new_json + content[end_bracket + 1:]

def leer_ixc_excel(path):
    """Lee el Excel de IXC: col A=Fecha(DD/MM/YYYY), col E=Total. Devuelve {YYYY-MM-DD: total}."""
    if not os.path.exists(path):
        print(f"  INFO IXC Excel no encontrado: {os.path.basename(path)} (opcional)")
        return {}
    print(f"  Leyendo IXC: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_raw, _, _, _, total = (row[i] if i < len(row) else None for i in range(5))
        if fecha_raw is None or total is None:
            continue
        # Parsear fecha
        iso = ""
        if isinstance(fecha_raw, (datetime, date)):
            iso = fecha_raw.strftime("%Y-%m-%d")
        elif isinstance(fecha_raw, str):
            m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", fecha_raw.strip())
            if m: iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        if iso and re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
            try:
                data[iso] = int(total)
            except (ValueError, TypeError):
                pass
    wb.close()
    print(f"  -> {len(data)} fechas IXC")
    return data

def es_sin_stock(r):
    """Filtra solo registros cuyo comentario u ops contienen 'sin stock'."""
    com = r.get("com", "").lower()
    ops = r.get("ops", "").lower()
    est = r.get("est", "").lower()
    return "sin stock" in com or "sin stock" in ops or "sin stock" in est

def reemplazar_objeto(content, variable, obj):
    """Reemplaza VAR = {...} en el HTML."""
    m = re.search(rf'(?<!\w){variable}\s*=\s*\{{', content)
    if not m:
        # No existe, insertar antes del primer uso
        return None
    start = m.end() - 1
    depth = 0; in_str = False; i = start
    while i < len(content):
        c = content[i]
        if in_str:
            if c == '\\': i += 2; continue
            elif c == '"': in_str = False
        else:
            if c == '"': in_str = True
            elif c == '{': depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0: end = i; break
        i += 1
    new_json = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    return content[:start] + new_json + content[end + 1:]

def actualizar_html(records_main, records_ss, ixc_data=None):
    if not os.path.exists(DASHBOARD):
        print(f"ERROR No se encontro: {DASHBOARD}")
        return False
    with open(DASHBOARD, "r", encoding="utf-8") as f:
        content = f.read()

    # -- Cifrar datos antes de embeber ----------------------------------------
    print(f"  Cifrando datos con clave configurada...")
    json_ss   = json.dumps(records_ss,    ensure_ascii=False, separators=(",", ":"))
    json_main = json.dumps(records_main,  ensure_ascii=False, separators=(",", ":"))
    json_ixc  = json.dumps(ixc_data or {}, ensure_ascii=False, separators=(",", ":"))

    enc_ss   = cifrar_datos(json_ss,   CLAVE_TABLERO)
    enc_main = cifrar_datos(json_main, CLAVE_TABLERO)
    enc_ixc  = cifrar_datos(json_ixc,  CLAVE_TABLERO)
    print(f"  Cifrado OK: SS={len(enc_ss)} chars, MAIN={len(enc_main)} chars, IXC={len(enc_ixc)} chars")

    # -- Insertar cifrado en el HTML ------------------------------------------
    content = reemplazar_string(content, "RECORDS_SS_ENC", enc_ss)
    if content is None: return False
    print(f"OK RECORDS_SS_ENC: {len(records_ss)} registros cifrados")

    content = reemplazar_string(content, "RECORDS_ENC", enc_main)
    if content is None: return False
    print(f"OK RECORDS_ENC: {len(records_main)} registros cifrados")

    content = reemplazar_string(content, "_ixcData_ENC", enc_ixc)
    if content is None:
        print("  INFO _ixcData_ENC no encontrado (se usara vacio en el tablero)")
    else:
        print(f"OK _ixcData_ENC: {len(ixc_data or {})} fechas IXC cifradas")

    import os as _os
    with open(DASHBOARD, "w", encoding="utf-8") as f:
        f.write(content)
        f.flush()
        _os.fsync(f.fileno())
    return True

# -- Main ---------------------------------------------------------------------
if __name__ == "__main__":
    print(f"\n{'='*50}")
    print(f"  Tablero GDI - Generador de Dashboard")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}\n")

    records_main = []
    for path in EXCELS_MAIN:
        records_main.extend(leer_excel(path))
    print(f"\nRECORDS (principal): {len(records_main)} registros\n")

    records_ss_raw = []
    for path in EXCELS_SS:
        records_ss_raw.extend(leer_excel(path))
    records_ss = [r for r in records_ss_raw if es_sin_stock(r)]
    print(f"RECORDS_SS (sin stock): {len(records_ss)} de {len(records_ss_raw)} registros filtrados\n")

    if not records_main:
        print("WARNING Sin registros en archivo principal.")
        exit(1)

    ixc_data = leer_ixc_excel(IXC_EXCEL)

    ok = actualizar_html(records_main, records_ss, ixc_data)
    if not ok:
        exit(1)

    print("\nOK Dashboard actualizado.")
